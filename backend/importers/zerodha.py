import csv
import io
from datetime import date
from .base import CSVImporter, HoldingDTO, StockTradeDTO


def _parse_float(s) -> float:
    if isinstance(s, (int, float)):
        return float(s)
    return float(str(s).replace(",", "").strip() or "0")


class ZerodhaHoldingsImporter(CSVImporter):
    """Imports from Zerodha Console → Portfolio → Holdings.

    Accepts both CSV (kite.zerodha.com download) and XLSX (Console holdings statement).
    This is a position snapshot — no actual trade dates are available, so no
    StockTrade records are created. XIRR will be null until a tradebook is imported.

    CSV columns:  Symbol / Instrument, ISIN, Qty., Avg. cost, LTP
    XLSX columns: Symbol, ISIN, Sector, Quantity Available, Average Price, Previous Closing Price
    """

    async def parse(self, content: bytes, filename: str, password: str = "") -> list[HoldingDTO]:
        if filename.lower().endswith(".xlsx"):
            return self._parse_xlsx(content)
        return self._parse_csv(content)

    def _parse_csv(self, content: bytes) -> list[HoldingDTO]:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
        holdings = []
        for row in reader:
            symbol = (row.get("Symbol") or row.get("Instrument") or "").strip()
            if not symbol:
                continue
            qty = _parse_float(row.get("Qty.", "0"))
            avg_cost = _parse_float(row.get("Avg. cost", "0"))
            ltp = _parse_float(row.get("LTP", "0"))
            invested = round(qty * avg_cost, 2)
            current = round(qty * ltp, 2)
            holdings.append(HoldingDTO(
                instrument_type="stock",
                display_name=symbol,
                asset_class="equity",
                invested_amount=invested,
                current_value=current,
                metadata={
                    "symbol": f"{symbol}.NS",
                    "isin": row.get("ISIN", "").strip(),
                    "exchange": "NSE",
                    "quantity": qty,
                    "average_price": avg_cost,
                    "current_price": ltp,
                },
                confidence_score=0.95,
            ))
        return holdings

    def _parse_xlsx(self, content: bytes) -> list[HoldingDTO]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Prefer 'Equity' sheet; fall back to first sheet
        ws = wb["Equity"] if "Equity" in wb.sheetnames else wb.active

        # Scan rows to find the header row (contains 'Symbol')
        header_row_idx = None
        col_map: dict[str, int] = {}
        all_rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(all_rows):
            for j, cell in enumerate(row):
                if cell == "Symbol":
                    header_row_idx = i
                    # Build column index from this row
                    for k, h in enumerate(row):
                        if h is not None:
                            col_map[str(h).strip()] = k
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            return []

        def _col(row, name: str):
            idx = col_map.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        holdings = []
        for row in all_rows[header_row_idx + 1:]:
            symbol = _col(row, "Symbol")
            if not symbol:
                continue
            symbol = str(symbol).strip()
            isin = str(_col(row, "ISIN") or "").strip()
            sector = str(_col(row, "Sector") or "").strip()
            qty = _parse_float(_col(row, "Quantity Available") or 0)
            avg_price = _parse_float(_col(row, "Average Price") or 0)
            prev_close = _parse_float(_col(row, "Previous Closing Price") or 0)
            invested = round(qty * avg_price, 2)
            current = round(qty * prev_close, 2)

            holdings.append(HoldingDTO(
                instrument_type="stock",
                display_name=symbol,
                asset_class="equity",
                invested_amount=invested,
                current_value=current,
                metadata={
                    "symbol": f"{symbol}.NS",
                    "isin": isin,
                    "exchange": "NSE",
                    "sector": sector,
                    "quantity": qty,
                    "average_price": avg_price,
                    "current_price": prev_close,
                },
                confidence_score=0.95,
                warnings=["Current price is previous day's closing price — refresh for live price"],
            ))
        return holdings


class ZerodhaTradebookImporter(CSVImporter):
    """Imports from Zerodha Console → Reports → Tradebook.

    Accepts both CSV and XLSX. Each row is an individual trade execution.
    Emits StockTradeDTO objects into stock_trades (not the generic transactions table).
    The HoldingDTO carries the aggregated position derived from the trade history.

    CSV columns (lowercase):  symbol, trade_date, trade_type, quantity, price, trade_id
    CSV columns (uppercase):  Symbol, ISIN, Trade Type, Quantity, Price, Trade Date
    XLSX columns:             Symbol, ISIN, Trade Date, Exchange, Segment, Series,
                              Trade Type, Auction, Quantity, Price, Trade ID, Order ID, Order Execution Time
    """

    async def parse(self, content: bytes, filename: str, password: str = "") -> list[HoldingDTO]:
        if filename.lower().endswith(".xlsx"):
            return self._parse_xlsx(content)
        return self._parse_csv(content)

    def _parse_csv(self, content: bytes) -> list[HoldingDTO]:
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
        by_symbol: dict[str, list] = {}
        for row in reader:
            symbol = (row.get("symbol") or row.get("Symbol") or "").strip()
            if not symbol:
                continue
            by_symbol.setdefault(symbol, []).append(row)

        return self._build_holdings(
            symbol_rows=[
                (symbol, [
                    {
                        "trade_type": (r.get("trade_type") or r.get("Trade Type") or "").upper(),
                        "quantity": _parse_float(r.get("quantity") or r.get("Quantity") or "0"),
                        "price": _parse_float(r.get("price") or r.get("Price") or "0"),
                        "trade_id": (r.get("trade_id") or r.get("Trade ID") or "").strip(),
                        "trade_date": (r.get("trade_date") or r.get("Trade Date") or "").strip()[:10],
                        "exchange": "NSE",
                        "segment": "EQ",
                        "isin": (r.get("isin") or r.get("ISIN") or "").strip(),
                    }
                    for r in rows
                ])
                for symbol, rows in by_symbol.items()
            ]
        )

    def _parse_xlsx(self, content: bytes) -> list[HoldingDTO]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Equity"] if "Equity" in wb.sheetnames else wb.active

        # Find header row and build column map
        all_rows = list(ws.iter_rows(values_only=True))
        header_row_idx = None
        col_map: dict[str, int] = {}
        for i, row in enumerate(all_rows):
            for j, cell in enumerate(row):
                if cell == "Symbol":
                    header_row_idx = i
                    for k, h in enumerate(row):
                        if h is not None:
                            col_map[str(h).strip()] = k
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            return []

        def _col(row, name: str):
            idx = col_map.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        by_symbol: dict[str, list] = {}
        for row in all_rows[header_row_idx + 1:]:
            symbol = _col(row, "Symbol")
            if not symbol:
                continue
            symbol = str(symbol).strip()
            raw_date = str(_col(row, "Trade Date") or "").strip()[:10]
            exchange = str(_col(row, "Exchange") or "NSE").strip().upper()
            segment = str(_col(row, "Segment") or "EQ").strip().upper()
            by_symbol.setdefault(symbol, []).append({
                "trade_type": str(_col(row, "Trade Type") or "").upper(),
                "quantity": _parse_float(_col(row, "Quantity") or 0),
                "price": _parse_float(_col(row, "Price") or 0),
                "trade_id": str(_col(row, "Trade ID") or "").strip(),
                "trade_date": raw_date,
                "exchange": exchange,
                "segment": segment,
                "isin": str(_col(row, "ISIN") or "").strip(),
            })

        return self._build_holdings(list(by_symbol.items()))

    def _build_holdings(self, symbol_rows: list[tuple[str, list[dict]]]) -> list[HoldingDTO]:
        holdings = []
        for symbol, trades_data in symbol_rows:
            stock_trades: list[StockTradeDTO] = []
            total_qty = 0.0
            total_cost = 0.0
            isin = ""
            exchange = "NSE"

            for t in trades_data:
                try:
                    tx_date = date.fromisoformat(t["trade_date"])
                except (ValueError, KeyError):
                    tx_date = date.today()

                is_buy = t["trade_type"] == "BUY"
                qty = t["quantity"]
                price = t["price"]
                amount = round(qty * price, 2)
                exchange = t.get("exchange", "NSE")
                if t.get("isin"):
                    isin = t["isin"]

                stock_trades.append(StockTradeDTO(
                    trade_date=tx_date,
                    trade_type="buy" if is_buy else "sell",
                    quantity=qty,
                    price=price,
                    amount=amount if is_buy else -amount,
                    exchange=exchange,
                    segment=t.get("segment", "EQ"),
                    trade_id=t.get("trade_id", ""),
                ))

                if is_buy:
                    total_qty += qty
                    total_cost += amount
                else:
                    total_qty -= qty

            # NSE → .NS suffix, BSE → .BO suffix for yfinance
            suffix = ".BO" if exchange == "BSE" else ".NS"
            holdings.append(HoldingDTO(
                instrument_type="stock",
                display_name=symbol,
                asset_class="equity",
                invested_amount=round(total_cost, 2),
                current_value=round(total_cost, 2),  # refreshed after import
                metadata={
                    "symbol": f"{symbol}{suffix}",
                    "isin": isin,
                    "exchange": exchange,
                    "quantity": max(total_qty, 0),
                },
                stock_trades=stock_trades,
                confidence_score=0.85,
                warnings=["Current price not available in tradebook — will be refreshed"],
            ))
        return [h for h in holdings if h.stock_trades]
