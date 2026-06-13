from .base import CSVImporter, HoldingDTO
from .zerodha import ZerodhaHoldingsImporter, ZerodhaTradebookImporter
from .groww import GrowwStocksImporter, GrowwMFImporter
from .kuvera import KuveraImporter
from .cams_cas import CAMSCASImporter
from .generic import GenericClaudeImporter

# XLSX magic bytes (ZIP/PK signature)
_XLSX_MAGIC = b"PK\x03\x04"


def detect_format(headers: list[str], file_preview: str) -> str:
    """Detect CSV format from headers and first-500-bytes preview."""
    header_set = set(headers)

    if "Consolidated Account Statement" in file_preview or "KFin Technologies" in file_preview:
        return "cams_cas"
    if {"Qty.", "Avg. cost", "LTP"}.issubset(header_set) and header_set & {"Symbol", "Instrument"}:
        return "zerodha_holdings"
    if {"Symbol", "ISIN", "Trade Type", "Quantity", "Price", "Trade Date"}.issubset(header_set):
        return "zerodha_tradebook"
    if {"symbol", "trade_date", "trade_type", "quantity", "price"}.issubset(header_set):
        return "zerodha_tradebook"
    if {"Stock Name", "Quantity", "Average Price", "Current Price", "Invested Value"}.issubset(header_set):
        return "groww_stocks"
    if {"Fund Name", "Units", "Average NAV", "Current NAV", "Invested Amount"}.issubset(header_set):
        return "groww_mf"
    if {"Scheme Name", "Folio", "Units", "Average NAV", "Current NAV"}.issubset(header_set):
        return "kuvera"
    return "generic"


def detect_xlsx_format(content: bytes) -> str:
    """Detect format of an XLSX file by peeking at sheet names and header row content."""
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheets = set(wb.sheetnames)

        # Zerodha Console holdings export has both 'Equity' and 'Mutual Funds' sheets
        if "Equity" in sheets and "Mutual Funds" in sheets:
            return "zerodha_holdings"

        # Zerodha tradebook has only 'Equity'; distinguish by scanning for header keywords
        if "Equity" in sheets:
            ws = wb["Equity"]
            for row in ws.iter_rows(values_only=True):
                row_vals = {str(v).strip() for v in row if v is not None}
                if "Trade Date" in row_vals:
                    return "zerodha_tradebook"
                if "Quantity Available" in row_vals:
                    return "zerodha_holdings"
    except Exception:
        pass
    return "generic"


def get_importer(fmt: str) -> CSVImporter:
    mapping = {
        "zerodha_holdings": ZerodhaHoldingsImporter(),
        "zerodha_tradebook": ZerodhaTradebookImporter(),
        "groww_stocks": GrowwStocksImporter(),
        "groww_mf": GrowwMFImporter(),
        "kuvera": KuveraImporter(),
        "cams_cas": CAMSCASImporter(),
        "generic": GenericClaudeImporter(),
    }
    return mapping.get(fmt, GenericClaudeImporter())
